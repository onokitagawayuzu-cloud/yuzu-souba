# -*- coding: utf-8 -*-
"""東京・大阪・京都・名古屋のゆず月次データ(数量・金額・平均単価)を取得する。

出力: data/monthly.json
  {"tokyo": {"2026-05": {"qty": .., "amount": .., "price": .., "origins": {..}}, ..},
   "osaka": {...}, "kyoto": {...}, "nagoya": {...}}

データ源:
  東京   市場統計情報明細データ(月次xlsx)  https://www.shijou.metro.tokyo.lg.jp/torihiki/geppo/
  大阪   月報CGIのCSV(本場+東部, 産地別)   https://www.shijou.city.osaka.jp/sikyo/tuki/hinmoku.html
  京都   月報青果部 品目別(野菜)xlsx       https://www.city.kyoto.lg.jp (第一市場)
  名古屋 月別取扱高 品目別xlsx(本場+北部)  https://www.city.nagoya.jp
"""
import csv
import io
import re
import sys
import tempfile
import warnings
from datetime import date
from pathlib import Path

import openpyxl

from common import DATA_DIR, http_get, http_get_text, load_json, save_json, to_num

warnings.filterwarnings("ignore")

OUT = DATA_DIR / "monthly.json"
START_YM = (2024, 1)  # ここから遡って取得


def month_range():
    today = date.today()
    y, m = START_YM
    while (y, m) <= (today.year, today.month):
        yield f"{y:04d}-{m:02d}"
        m += 1
        if m > 12:
            y, m = y + 1, 1


def load_xlsx_from_bytes(b):
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.write(b)
    tmp.close()
    wb = openpyxl.load_workbook(tmp.name, read_only=True, data_only=True)
    return wb, tmp.name


# ---------------------------------------------------------------- 東京
def fetch_tokyo(data):
    page = http_get_text("https://www.shijou.metro.tokyo.lg.jp/torihiki/geppo/",
                         encoding="utf-8")
    links = {}  # "2026-05" -> URL
    for m in re.finditer(r'href="(/documents/d/shijou/y(\d{6})meisai[^"]*)"', page):
        ym = f"{m.group(2)[:4]}-{m.group(2)[4:6]}"
        links.setdefault(ym, "https://www.shijou.metro.tokyo.lg.jp" + m.group(1))
    for ym in month_range():
        if ym in data and data[ym].get("qty"):
            continue
        url = links.get(ym)
        if not url:
            continue
        b = http_get(url)
        if b is None:
            continue
        wb, tmpname = load_xlsx_from_bytes(b)
        ws = wb.active
        qty = amount = 0
        origins = {}
        for row in ws.iter_rows(values_only=True):
            if not row or len(row) < 17:
                continue
            if row[11] == "ゆず":
                q, a = row[15] or 0, row[16] or 0
                qty += q
                amount += a
                o = origins.setdefault(str(row[13]), {"qty": 0, "amount": 0})
                o["qty"] += q
                o["amount"] += a
        wb.close()
        Path(tmpname).unlink(missing_ok=True)
        if qty:
            for o in origins.values():
                o["price"] = round(o["amount"] / o["qty"]) if o["qty"] else None
            data[ym] = {"qty": qty, "amount": amount,
                        "price": round(amount / qty), "origins": origins}
            print(f"  東京 {ym}: {qty:,}kg {data[ym]['price']}円/kg")


# ---------------------------------------------------------------- 大阪
def fetch_osaka(data):
    base = "https://www.shijou.city.osaka.jp/data/webroot/tuki/hinmoku"
    for ym in month_range():
        if ym in data and data[ym].get("qty"):
            continue
        y, m = ym.split("-")
        qty = amount = 0
        origins = {}
        found = False
        for icode in ("1", "2"):  # 本場, 東部
            url = f"{base}/{y}/{y}{m}-{icode}1-1-39300000.csv"
            text = http_get_text(url)
            if text is None or "ゆず" not in text:
                continue
            found = True
            rows = list(csv.reader(io.StringIO(text)))
            in_table = False
            for r in rows:
                if not r or not r[0].strip():
                    continue
                c0 = r[0].strip()
                if c0.startswith("産"):
                    in_table = True
                    continue
                if in_table:
                    if c0.startswith(("比率", "※", "【", "★", "合")):
                        if c0.startswith("合"):
                            continue
                        break
                    q, a = to_num(r[1]), to_num(r[4])
                    if q is None:
                        continue
                    qty += q
                    amount += a or 0
                    o = origins.setdefault(c0, {"qty": 0, "amount": 0})
                    o["qty"] += q
                    o["amount"] += a or 0
        if found and qty:
            for o in origins.values():
                o["price"] = round(o["amount"] / o["qty"]) if o["qty"] else None
            data[ym] = {"qty": qty, "amount": amount,
                        "price": round(amount / qty), "origins": origins}
            print(f"  大阪 {ym}: {qty:,}kg {data[ym]['price']}円/kg")


# ---------------------------------------------------------------- 京都
KYOTO_YEAR_PAGES = {
    2024: "https://www.city.kyoto.lg.jp/sankan/page/0000322250.html",
    2025: "https://www.city.kyoto.lg.jp/sankan/page/0000339440.html",
    2026: "https://www.city.kyoto.lg.jp/sankan/page/0000350833.html",
}


def fetch_kyoto(data):
    for year, ypage in KYOTO_YEAR_PAGES.items():
        need = [ym for ym in month_range()
                if ym.startswith(str(year)) and not data.get(ym, {}).get("qty")]
        if not need:
            continue
        html = http_get_text(ypage, encoding="utf-8")
        if html is None:
            continue
        # 月報青果部(令和X年Y月) -> 月ページURL
        month_links = {}
        for m in re.finditer(
                r'<a[^>]*href="([^"]+)"[^>]*>[^<]*青果部[^<]*?(\d{1,2})月[^<]*</a>', html):
            month_links[int(m.group(2))] = m.group(1)
        for ym in need:
            mon = int(ym.split("-")[1])
            if mon not in month_links:
                continue
            murl = month_links[mon]
            if murl.startswith("/"):
                murl = "https://www.city.kyoto.lg.jp" + murl
            mhtml = http_get_text(murl, encoding="utf-8")
            if mhtml is None:
                continue
            fm = re.search(r'href="([^"]*hinmokuyasai[^"]*\.xlsx?)"', mhtml)
            if not fm:
                continue
            furl = fm.group(1)
            if furl.startswith("./"):
                # ページは /sankan/page/xxx.html だが実ファイルは /sankan/cmsfiles/ 配下
                furl = murl.rsplit("/", 1)[0] + furl[1:]
            elif furl.startswith("/"):
                furl = "https://www.city.kyoto.lg.jp" + furl
            b = http_get(furl)
            if b is None and "/page/cmsfiles/" in furl:
                b = http_get(furl.replace("/page/cmsfiles/", "/cmsfiles/"))
            if b is None:
                continue
            wb, tmpname = load_xlsx_from_bytes(b)
            found = None
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    if row and isinstance(row[0], str) and re.match(r"^\d*ゆず", row[0].strip()):
                        vals = [c for c in row[1:] if c is not None]
                        if len(vals) >= 3:
                            found = vals[:3]  # 数量, 金額, 平均単価
                            break
                if found:
                    break
            wb.close()
            Path(tmpname).unlink(missing_ok=True)
            if found:
                qty, amount, price = found
                data[ym] = {"qty": qty, "amount": amount, "price": round(price)}
                print(f"  京都 {ym}: {qty:,}kg {data[ym]['price']}円/kg")


# ---------------------------------------------------------------- 名古屋
NAGOYA_YEAR_PAGES = {
    2024: "https://www.city.nagoya.jp/kurashi/shisetsu/1037173/1016131/1034822/1016141/1034825/1016175/index.html",
    2025: "https://www.city.nagoya.jp/kurashi/shisetsu/1037173/1016131/1034822/1016141/1034825/1034826/index.html",
    2026: "https://www.city.nagoya.jp/kurashi/shisetsu/1037173/1016131/1034822/1016141/1034825/1046051/index.html",
}
WAREKI = {2024: "令和6年", 2025: "令和7年", 2026: "令和8年"}


def fetch_nagoya(data):
    for year, ypage in NAGOYA_YEAR_PAGES.items():
        need = [ym for ym in month_range()
                if ym.startswith(str(year)) and not data.get(ym, {}).get("qty")]
        if not need:
            continue
        html = http_get_text(ypage, encoding="utf-8")
        if html is None:
            continue
        month_links = {}
        wareki = WAREKI[year]
        for m in re.finditer(
                r'<a[^>]*href="([^"]+)"[^>]*>\s*' + wareki + r'(\d{1,2})月[^<]*</a>', html):
            month_links[int(m.group(2))] = m.group(1)
        for ym in need:
            mon = int(ym.split("-")[1])
            if mon not in month_links:
                continue
            murl = month_links[mon]
            if not murl.startswith("http"):
                murl = "https://www.city.nagoya.jp/" + murl.lstrip("./").replace("../", "")
            mhtml = http_get_text(murl, encoding="utf-8")
            if mhtml is None:
                continue
            fm = re.search(r'href="([^"]*_sei_hinmokbet\.xlsx?)"', mhtml)
            if not fm:
                continue
            furl = fm.group(1)
            if not furl.startswith("http"):
                furl = "https://www.city.nagoya.jp/" + furl.lstrip("./").replace("../", "")
            b = http_get(furl)
            if b is None:
                continue
            wb, tmpname = load_xlsx_from_bytes(b)
            found = None
            for ws in wb.worksheets:
                if "野菜" not in ws.title:
                    continue
                for row in ws.iter_rows(values_only=True):
                    if row and any(isinstance(c, str) and c.strip() == "ゆず類" for c in row):
                        vals = [c for c in row if isinstance(c, (int, float))]
                        if len(vals) >= 3:
                            found = vals[:3]  # 合計の数量, 金額, 平均単価
                            break
                if found:
                    break
            wb.close()
            Path(tmpname).unlink(missing_ok=True)
            if found:
                qty, amount, price = found
                data[ym] = {"qty": qty, "amount": amount, "price": round(price)}
                print(f"  名古屋 {ym}: {qty:,}kg {data[ym]['price']}円/kg")


def main():
    all_data = load_json(OUT)
    # 名古屋は2026-07に取得停止(fetch_nagoya関数は残置。過去分データはmonthly.jsonに保存済み)
    for key, fn in [("tokyo", fetch_tokyo), ("osaka", fetch_osaka),
                    ("kyoto", fetch_kyoto)]:
        all_data.setdefault(key, {})
        try:
            fn(all_data[key])
        except Exception as e:
            print(f"  !! {key} の取得でエラー: {e}")
        save_json(OUT, all_data)
    print("月次データ更新完了")


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
